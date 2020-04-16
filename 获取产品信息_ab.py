# -*- coding: utf-8 -*-
"""
Created on Tue Apr 14 09:36:49 2020

@author: sds
"""
import re
import requests
from bs4 import BeautifulSoup
import openpyxl
import time


headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/75.0.3770.100 Safari/537.36'}
'''
headers={
    "Accept":"text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Encoding": "gzip, deflate, br",
    "Accept-Language": "zh-CN,zh;q=0.8,en-US;q=0.5,en;q=0.3",
    "Connection":"keep-alive",
    "Host":    "36kr.com/newsflashes",
    "Upgrade-Insecure-Requests":"1",
    "User-Agent":"Mozilla/5.0 (Macintosh; Intel Mac OS X 10.13; rv:55.0) Gecko/20100101 Firefox/55.0"
}
'''

def get_html(url):
    resp = requests.get(url,  headers=headers)
    resp.encoding = resp.apparent_encoding
    html = resp.text
    #print(html)
    return html

def result(content):
    data_catalog=content[0].attrs.get('data-catalog')#型号
    data_date=content[0].attrs.get('data-date')#停产时间
    data_id=content[0].attrs.get('data-id')#产品状态
    productDescription=content[0].find_all('span',id="productDescription")[0].text#产品说明
    if data_id=='active':
        Replacement_Product=''
    else:
        try:
            Replacement_Product=re.findall(r'Replacement Product</div>(.*?)</div>', str(content[0]))[0]#替代产品
            #print(Replacement_Product[0])
        except:
            Replacement_Product=''
    return data_catalog,productDescription,data_id,data_date,Replacement_Product


if __name__ == '__main__':
    filename='物料编码-abplc.xlsx'
    wb = openpyxl.load_workbook(filename)
    sh = wb['Sheet1']
    print(sh.max_row)    
    for i in range(2,sh.max_row+1):
        d = sh.cell(row=i, column=4).value
        print('-----------------')
        print('正在查询第%d个' %i)
        print('型号：',d)
        e=re.sub("-", "",d)
        url="https://www.rockwellautomation.com.cn/global/support/product-compatibility-migration/lifecycle-status/results.page?productid=%s" %e
        html_content = get_html(url)
        soup = BeautifulSoup(html_content,'html.parser',from_encoding='utf-8')
        content0 =soup.find_all('h4',class_="error")
        content1 =soup.find_all('div',class_="row lifecycle-results active %s" %d)
        content2 =soup.find_all('div',class_="row lifecycle-results mature %s" %d)
        content3 =soup.find_all('div',class_="row lifecycle-results endOfLife %s" %d)
        content4 =soup.find_all('div',class_="row lifecycle-results discontinued %s" %d)
        content=content1+content2+content3+content4
        if len(content0)!=0 or len(content)==0 :
            sh.cell(row=i,column=5).value='未查到'
            print('未查到')
        else:
            res=result(content)
            for j in range(len(res)) :
                print(res[j])
                sh.cell(row=i,column=5+j).value=result(content)[j]
        wb.save(filename)
        time.sleep(3)
    #print(content)  


